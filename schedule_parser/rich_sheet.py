"""Загрузка таблиц с форматированием (цвет ячейки, жирный шрифт).

Google Sheets публикует XLSX-экспорт, в котором сохраняются:
- цвет заливки ячейки (определяет тип занятия: лекция/практика/лаба/...);
- жирный шрифт фрагментов внутри ячейки (название предмета всегда жирное);
- объединённые ячейки (одно занятие на 2 пары → rowspan=2).

CSV-экспорт всю эту информацию теряет, поэтому он остаётся только как fallback
для офлайн-тестов на старых фикстурах.
"""
from __future__ import annotations

import io
import urllib.request
from dataclasses import dataclass, field
from typing import Iterable


@dataclass
class CellRun:
    """Один фрагмент текста в ячейке с признаком жирности."""
    text: str
    bold: bool = False


@dataclass
class RichCell:
    """Ячейка таблицы с метаданными форматирования.

    Для CSV-источника runs состоит из одного фрагмента с bold=False, fill=None,
    rowspan=colspan=1.
    """
    value: str = ""
    runs: list[CellRun] = field(default_factory=list)
    fill: str | None = None  # hex без '#', напр. 'FFD9EAD3'
    rowspan: int = 1
    colspan: int = 1

    @property
    def text(self) -> str:
        return self.value

    @property
    def bold_text(self) -> str:
        """Конкатенация всех жирных фрагментов (со сохранением переводов строк)."""
        return "".join(r.text for r in self.runs if r.bold)

    def lines_with_bold(self) -> list[tuple[str, bool]]:
        """Разбить содержимое ячейки на строки с дробной долей жирного текста.

        Каждая строка получает флаг `is_bold = (доля жирных символов > 0.5)`.
        """
        out: list[tuple[str, bool]] = []
        buf = ""
        bold_chars = 0
        total = 0
        for run in self.runs:
            for ch in run.text:
                if ch == "\n":
                    if buf.strip() or total > 0:
                        out.append((buf, total > 0 and bold_chars / total > 0.5))
                    buf = ""
                    bold_chars = 0
                    total = 0
                else:
                    buf += ch
                    total += 1
                    if run.bold:
                        bold_chars += 1
        if total > 0 or buf.strip():
            out.append((buf, total > 0 and bold_chars / total > 0.5))
        return out


@dataclass
class RichRow:
    """Строка таблицы — список RichCell."""
    cells: list[RichCell] = field(default_factory=list)

    def value(self, col: int) -> str:
        if 0 <= col < len(self.cells):
            return self.cells[col].value
        return ""

    def cell(self, col: int) -> RichCell:
        if 0 <= col < len(self.cells):
            return self.cells[col]
        return RichCell()

    def __len__(self) -> int:
        return len(self.cells)


def _stringify(value: object) -> str:
    """Целые числа сохраняем как '117', а не '117.0' (Google Sheets любит float)."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def xlsx_bytes_to_rich_rows(data: bytes) -> list[RichRow]:
    """Распарсить XLSX-байты (один лист) в список RichRow с сохранением
    цветов заливки, жирного текста и объединённых ячеек.
    """
    import openpyxl  # ленивая зависимость
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, rich_text=True)
    ws = wb.active

    # Карта rowspan/colspan для левых-верхних ячеек объединённых диапазонов.
    rowspan: dict[tuple[int, int], int] = {}
    colspan: dict[tuple[int, int], int] = {}
    for mr in ws.merged_cells.ranges:
        rowspan[(mr.min_row, mr.min_col)] = mr.max_row - mr.min_row + 1
        colspan[(mr.min_row, mr.min_col)] = mr.max_col - mr.min_col + 1

    rows: list[RichRow] = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    for r in range(1, max_row + 1):
        cells: list[RichCell] = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            raw = cell.value

            runs: list[CellRun] = []
            if isinstance(raw, CellRichText):
                full_parts: list[str] = []
                for chunk in raw:
                    if isinstance(chunk, TextBlock):
                        is_bold = bool(getattr(chunk.font, "b", False))
                        runs.append(CellRun(text=chunk.text, bold=is_bold))
                        full_parts.append(chunk.text)
                    else:
                        runs.append(CellRun(text=str(chunk), bold=False))
                        full_parts.append(str(chunk))
                value = "".join(full_parts)
            else:
                value = _stringify(raw)
                # Если у ячейки cell-level bold (без рантаймов), считаем весь текст жирным.
                cell_bold = bool(cell.font and cell.font.bold) if value else False
                runs = [CellRun(text=value, bold=cell_bold)] if value else []

            fill: str | None = None
            try:
                fg = cell.fill.fgColor if cell.fill else None
                rgb = getattr(fg, "rgb", None) if fg is not None else None
                if isinstance(rgb, str):
                    fill = rgb
            except Exception:
                fill = None

            cells.append(RichCell(
                value=value,
                runs=runs,
                fill=fill,
                rowspan=rowspan.get((r, c), 1),
                colspan=colspan.get((r, c), 1),
            ))
        rows.append(RichRow(cells=cells))
    return rows


def fetch_xlsx_rich_rows(sheet_id: str, gid: str) -> list[RichRow]:
    """Скачать XLSX-экспорт листа из Google Sheets и распарсить в RichRow."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid={gid}"
    with urllib.request.urlopen(url, timeout=30) as resp:
        data = resp.read()
    return xlsx_bytes_to_rich_rows(data)


def read_local_xlsx(path: str) -> list[RichRow]:
    with open(path, "rb") as fh:
        return xlsx_bytes_to_rich_rows(fh.read())


def csv_rows_to_rich_rows(rows: Iterable[list[str]]) -> list[RichRow]:
    """Адаптер для CSV-источника: создаёт RichRow без форматирования."""
    out: list[RichRow] = []
    for row in rows:
        cells = [
            RichCell(
                value=v or "",
                runs=[CellRun(text=v or "", bold=False)] if v else [],
                fill=None,
                rowspan=1,
                colspan=1,
            )
            for v in row
        ]
        out.append(RichRow(cells=cells))
    return out
